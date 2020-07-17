import openpyxl
import requests
import sys
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import *
from openpyxl.chart import Reference, PieChart
from functions.functionLibrary import *

"""
    Week 5 - Spreadsheets Assignment
    Excel Spreadsheet Assignment
    Victor Ifezue
    Professor Tom Petz 
    
    I did all the average calculation with a 
    method instead of Excel Formulas
       
"""

sourceURL = "https://ool-content.walshcollege.edu/CourseFiles/IT/IT414/MASTER/Week05/WI20-Assignment/sales_data.html"
request = requests.get(sourceURL)
my_workbook = openpyxl.Workbook()
curr_sheet = my_workbook.active
curr_sheet.title = "My WorkBook"
my_workbook.save("files/my_workbook.xlsx")

# check for a sucessful response or exit app
if request.status_code != 200:
    exit()

requestText = request.text
beautifulSoupText = BeautifulSoup(requestText, "html.parser")
my_data = beautifulSoupText.findAll("tr")

# clean the dataset and return a clean array with values
cleanDataSet = []
counter = 0
while counter < len(my_data):
    rowValue = my_data[counter].getText()
    rowValue = str(rowValue).split('\n')
    while("" in rowValue):
        rowValue.remove("")
        if '  ' in rowValue:
            rowValue.remove('  ')
    cleanDataSet.append(rowValue)
    counter = counter + 1


# grab distinct products and put in list
productList = []
for item in cleanDataSet:
    if item[3] != "Item":
        if item[3] not in productList:
            productList.append(item[3])

# grab product and units sold and put in dictionary
productCount = []
for productItem in cleanDataSet:
    if productItem[3] in productList:
        value = {"product": productItem[3], "quantity": productItem[4]}
        productCount.append(value)

# calculate the average products sold by product
averageProductCount = []
for productValue in productList:
    number = []
    for item in productCount:
        if productValue == item["product"]:
            number.append(item["quantity"])
    average = calculateAverage(number)
    averageProductCount.append([productValue, int(average)])

# TITLE
curr_sheet.merge_cells("A1:D1")
curr_sheet["A1"] = "Sales Data"
curr_sheet["A1"].font = Font(sz=20.0, b=True, color="33E8FF")
curr_sheet["A1"].alignment = Alignment(horizontal="center")
my_workbook.save("files/my_workbook.xlsx")

# Import the Clean Product Data
row_count = 2
for item in cleanDataSet:

    col_count = 1
    curr_sheet.cell(row=row_count, column=col_count).value = item[3]
    col_count = col_count + 1
    curr_sheet.cell(row=row_count, column=col_count).value = item[4]
    row_count = row_count + 1
    col_letter = get_column_letter(col_count)
my_workbook.save("files/my_workbook.xlsx")

# Import the Clean Product Data Averages
row_count = row_count
for item in averageProductCount:
    col_count = 1
    curr_sheet.cell(
        row=row_count, column=col_count).value = item[0] + " Average:"
    col_count = col_count + 1
    curr_sheet.cell(row=row_count, column=col_count).value = item[1]
    row_count = row_count + 1
my_workbook.save("files/my_workbook.xlsx")
